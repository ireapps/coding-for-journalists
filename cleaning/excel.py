#! usr/bin/python

from openpyxl import load_workbook
import shutil

# We're going to be messing with an existing file, so let's clone our Excel
# file just to be on the safe side.

shutil.copyfile('Candidates.xlsx', 'Candidates_WORKING.xlsx')

# Load our Excel workbook into memory
loc = 'Candidates_WORKING.xlsx'
wb = load_workbook(filename = loc)

# List the sheets we find in the workbook
for sheet in wb.sheetnames:
	print sheet

# Active sheet will default to the first; we can also select it
sheet_read = wb.get_sheet_by_name('Sheet1')

# We can iterate through rows an do things based on the pattern that
# we find. First, though, we need to decide where to put the data.
# Why not another blank sheet in the same book?

sheet_write = wb.get_sheet_by_name('Sheet2')

# Right now there's nothing on the sheet; the max row and column would be '1'
sheet_write.max_row
sheet_write.max_column

# Max_row can help us keep track of where we are as we write to the book.
# Let's start by writing a header for all the information we're going to
# collect.
column_names = ['Race', 'Candidate Name', 'Filing Date', 'Street Address', 'City', 'State', 'Zip', 'Phone Number', 'Email Address']

for header in column_names:
	sheet_write.cell(column = column_names.index(header) + 1, row = 1, value = header)

# This is our first modification to the workbook; let's save our changes.
wb.save(filename = loc)

# We need something outside the loop to hold onto the race name

race_holder = ''

# Let's walk through each row in our sheet with .rows
for ws_row in sheet_read.rows:
	# Headers are duplicated for each race in this sheet; we can grab the
	# race name in the cell directly above and hold onto it IF we encounter
	# that first header item, "Candidate's Name."
	if ws_row[0].value == "Candidate's Name":
		cell_above = ws_row[0].column + str(ws_row[0].row - 1)
		race_holder = sheet_read[cell_above].value
	# The other condition to hunt for: Candidate info is spread out over two
	# rows. If there's something in the first and second columns, that means
	# it's a candidate, and we can go to work on that row and the row below it.
	elif ws_row[0].value != None and ws_row[1].value != None:
		row_below = ws_row[0].row + 1
		cand_name = ws_row[0].value
		st_addr = ws_row[1].value
		phone = ws_row[3].value
		filed_date = ws_row[4].value
		city_st_zip = sheet_read['B'+str(row_below)].value
		# Let's break apart the combined city, state and zip
		city = city_st_zip.split(',')[0]
		st_zip = city_st_zip.split(',')[1].split()
		st = st_zip[0]
		zip = st_zip[1]
		email = sheet_read['D'+str(row_below)].value
		# Let's write the row to Sheet2, which we've already queued up. Watch
		# out for the date, which we'll reformat from a Python date object.
		row = [race_holder, cand_name, filed_date.strftime('%m/%d/%Y'), st_addr, city, st, zip, phone, email]
		write_row = sheet_write.max_row + 1
		for item in row:
			sheet_write.cell(column = row.index(item) + 1, row = write_row, value = item)

# Save changes to the file.
wb.save(filename = loc)
		