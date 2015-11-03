# Excel still has a place in your work life with Python. People have written
# libraries so that you can read and write the modern XML-based file format
# for Microsoft Excel. Note: xlrd and xlwt exist for older Excel formats (.xls)

from openpyxl import Workbook

# Let's create an Excel workbook.
wb = Workbook()

# It starts with a default sheet (called 'Sheet') but let's make another.
wb.create_sheet('MyStuff')

# Print the names of the sheets in this new, unsaved workbook.
print wb.sheetnames

# Set this new sheet to be the active sheet.
wb.active = 1

# Print the name of the active sheet.
print wb.active

# We'll select this new sheet (again) and start adding some basic data to it.
ws = wb.get_sheet_by_name('MyStuff')

# If we dislike the title, we can change it.
ws.title = 'OtherStuff'

# Let's assign the first three columns of the workbook some header names.
ws['A1'] = 'Name'
ws['B1'] = 'Company'
ws['C1'] = 'Salary'

# We can retrieve values in cells in a similar manner.
ws['B1'].value

# Let's make a list of values to put into the second row, right underneath.
person = ['Laura Green', 'Dynamic Dynamics', '12550000']

for detail in person:
    ws.cell(column=person.index(detail)+1, row=2, value=detail)

# We can also comb through rows and get the output.
for row in ws.rows[1:]:
    print ', '.join([row[0].value, row[1].value, row[2].value])

# One catch: it's basically all in memory until we write the workbook to disk.
wb.save('mytest.xlsx')
